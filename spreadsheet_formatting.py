# this file contains functions involving the openpyxl package
# isolated from other code files because this could get messy

import openpyxl
from colour import Color
import os
import re


def name_colors():
    # outputs a dict of colors, all hex codes keyed by name
    # for now, just the one's used last year
    color = dict()
    color["Red"] = "FF0000"
    color["Green"] = "00FF00"
    color["Blue"] = "0000FF"
    color["Orange"] = "FF5500"
    color["Purple"] = "8000FF"
    color["Black"] = "000000"
    color["White"] = "FFFFFF"
    color["Yellow"] = "CCCC00"

    return color


def output_fantasy_results(dest_filename, title, results):
    # currently assuming results take format of final_results from fantasy_spreadsheet_io (i.e. a dict of tuples with
    # 2 dicts: starting players and relevant scores, and sub players with relevant scores
    # if we want to fill the sheet with all the stats used to find scores, we can expand that functionality later
    # additionally assuming players and teams take the form of scraped player and team data
    wb = openpyxl.Workbook()
    # dest_filename = 'fantasy_book.xlsx'

    ws1 = wb.active
    ws1.title = title

    # style setup below
    # colors for styling
    color = name_colors()
    # generic , oft-reused styles
    base_style = openpyxl.styles.NamedStyle(name="base_style")
    base_style.font = openpyxl.styles.Font(name='Arial', size=10)
    base_style.alignment = openpyxl.styles.Alignment(horizontal='left')

    bold_style = openpyxl.styles.NamedStyle(name="bold_style")
    bold_style.font = openpyxl.styles.Font(name='Arial', size=10, bold=True)
    bold_style.alignment = openpyxl.styles.Alignment(horizontal='center')

    # stat boxes have borders
    stat_style = openpyxl.styles.NamedStyle(name="stat_style")
    stat_style.font = openpyxl.styles.Font(name='Arial', size=10)
    stat_style.alignment = openpyxl.styles.Alignment(horizontal='right')
    stat_style.border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='thin'),
                                               right=openpyxl.styles.Side(border_style='thin'),
                                               top=openpyxl.styles.Side(border_style='thin'),
                                               bottom=openpyxl.styles.Side(border_style='thin'))

    # team stat boxes do NOT have borders
    tstat_style = openpyxl.styles.NamedStyle(name="tstat_style")
    tstat_style.font = openpyxl.styles.Font(name='Arial', size=10)
    tstat_style.alignment = openpyxl.styles.Alignment(horizontal='right')

    name_style = openpyxl.styles.NamedStyle(name="name_style")
    name_style.font = openpyxl.styles.Font(name='Arial', size=14, bold=True, italic=True, color=Color("White").get_hex_l()[1:])
    # cell backgrounds: must be defined in-loop due to reliance on players' chosen colors

    # column formatting so that the sheet looks nice
    ws1.row_dimensions[1].height = 30
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 15
    ws1.column_dimensions['H'].width = 15
    ws1.column_dimensions['I'].width = 15
    ws1.column_dimensions['J'].width = 15

    # the rows written to are relative to the number of players in the game
    # we can get these from looping through results
    row = 1
    # top row gives the name of the winner in a special message format
    # to get the winner, compare scores of all members of results, the key with highscore wins
    winner = list()
    win_color = list()
    for name in results:
        if len(winner) == 0 or results[name].get_score() > results[winner[0]].get_score():
            winner = [name]
            win_color = [results[name].color]
        elif results[name].get_score() == results[winner[0]].get_score():
            winner.append(name)
            win_color.append(results[name].color)
    # begin by merging top cells
    ws1.merge_cells('B'+str(row)+':F'+str(row))
    winner_cell = ws1['B'+str(row)]
    winner_string = "#"
    for i in range(len(winner)):
        winner_string += winner[i]
    winner_cell.value = winner_string +"Win"
    # special styling for the header
    winner_cell.alignment = openpyxl.styles.Alignment(horizontal = "center")
    if (len(win_color) == 1):
        winner_cell.font = openpyxl.styles.Font(name='Arial', size=24, bold=True, italic=True, color=Color(win_color[0]).get_hex_l()[1:])
    else:
        winner_cell.font = openpyxl.styles.Font(name='Arial', size=24, bold=True, italic=True, color=Color("Black").get_hex_l()[1:])

    row += 1
    for name in results:
        # print(name)
        ws1.merge_cells('A' + str(row) + ':H'+str(row))
        name_cell = ws1['A'+str(row)]
        name_cell.value = name
        # special styling for name cell
        name_cell.style = name_style
        name_cell.fill = openpyxl.styles.PatternFill(fgColor=Color(results[name].color).get_hex_l()[1:], fill_type='solid')
        # after name, 2 cells merged to present score
        ws1.merge_cells('I'+str(row) + ':J' + str(row))
        score_cell = ws1['I' + str(row)]
        # score_cell.value = results[name][4]
        score_cell.value = results[name].get_score()
        score_cell.style = name_style
        score_cell.alignment = openpyxl.styles.Alignment(horizontal='right')
        score_cell.fill = openpyxl.styles.PatternFill(fgColor=Color(results[name].color).get_hex_l()[1:], fill_type='solid')
        # space after name
        row += 1
        # light_color = Color(results[name][3])
        # light_color.set_luminance(0.90) # this section will be a lighter shade of the primary color
        # for col in 'ABCDEFGHIJ':
        #     ws1[col+str(row)].fill = openpyxl.styles.PatternFill(fgColor=light_color.get_hex_l()[1:], fill_type='solid')
        row += 1
        # fill in column headers for team stats
        ws1['C'+str(row)] = "Wins"
        ws1['D'+str(row)] = "Towers"
        ws1['E'+str(row)] = "Barons"
        ws1['F' + str(row)] = "Dragons"
        ws1['G' + str(row)] = "Rift Heralds"
        ws1['H' + str(row)] = "<30 Min Wins"
        ws1['J' + str(row)] = "Team Points"
        # apply formatting to each cell
        ws1['C' + str(row)].style = bold_style
        ws1['D' + str(row)].style = 'bold_style'
        ws1['E' + str(row)].style = 'bold_style'
        ws1['F' + str(row)].style = 'bold_style'
        ws1['G' + str(row)].style = 'bold_style'
        ws1['H' + str(row)].style = 'bold_style'
        ws1['J' + str(row)].style = 'bold_style'

        # apply background light color
        # for col in 'ABCDEFGHIJ':
        #     ws1[col+str(row)].fill = openpyxl.styles.PatternFill(fgColor=light_color.get_hex_l()[1:], fill_type='solid')
        row += 1

        ws1['A' + str(row)] = "Team"
        team = results[name].team
        ws1['B' + str(row)] = team.name
        # now, fill in team stats from teams and final score from results
        ws1['C' + str(row)] = team.wins
        ws1['D' + str(row)] = team.towers
        ws1['E' + str(row)] = team.barons
        ws1['F' + str(row)] = team.dragons
        ws1['G' + str(row)] = team.heralds
        ws1['H' + str(row)] = team.fastwins
        ws1['J' + str(row)] = team.get_score()
        # apply formatting to each cell
        ws1['A' + str(row)].style = 'bold_style'
        ws1['B' + str(row)].style = base_style
        ws1['C'+str(row)].style = tstat_style
        ws1['D' + str(row)].style = 'tstat_style'
        ws1['E' + str(row)].style = 'tstat_style'
        ws1['F' + str(row)].style = 'tstat_style'
        ws1['G' + str(row)].style = 'tstat_style'
        ws1['H' + str(row)].style = 'tstat_style'
        ws1['J' + str(row)].style = 'tstat_style'
        # apply background color
        # for col in 'ABCDEFGHIJ':
        #     ws1[col+str(row)].fill = openpyxl.styles.PatternFill(fgColor=light_color.get_hex_l()[1:], fill_type='solid')
        row += 1

        # player stat headers
        ws1['C' + str(row)] = "Kills"
        ws1['D' + str(row)] = "Deaths"
        ws1['E' + str(row)] = "Assists"
        ws1['F' + str(row)] = "CS"
        ws1['G' + str(row)] = "Totals"
        # apply formatting
        ws1['C' + str(row)].style = 'bold_style'
        ws1['D' + str(row)].style = 'bold_style'
        ws1['E' + str(row)].style = 'bold_style'
        ws1['F' + str(row)].style = 'bold_style'
        ws1['G' + str(row)].style = 'bold_style'

        row += 1

        # player entries for each position
        # individual stats from players, name and score from results
        ws1['A' + str(row)] = "Top"
        top = results[name].starters[1]
        ws1['B' + str(row)] = top.name
        ws1['C' + str(row)] = top.k
        ws1['D' + str(row)] = top.d
        ws1['E' + str(row)] = top.a
        ws1['F' + str(row)] = top.cs
        ws1['G' + str(row)] = top.get_score()
        # apply formatting
        ws1['A'+str(row)].style = 'bold_style'
        ws1['B' + str(row)].style = 'base_style'
        ws1['C' + str(row)].style = stat_style
        ws1['D' + str(row)].style = 'stat_style'
        ws1['E' + str(row)].style = 'stat_style'
        ws1['F' + str(row)].style = 'stat_style'
        ws1['G' + str(row)].style = 'stat_style'

        row += 1

        ws1['A' + str(row)] = "Jungle"
        jung = results[name].starters[2]
        ws1['B' + str(row)] = jung.name
        ws1['C' + str(row)] = jung.k
        ws1['D' + str(row)] = jung.d
        ws1['E' + str(row)] = jung.a
        ws1['F' + str(row)] = jung.cs
        ws1['G' + str(row)] = jung.get_score()
        # apply formatting
        ws1['A' + str(row)].style = 'bold_style'
        ws1['B' + str(row)].style = 'base_style'
        ws1['C' + str(row)].style = 'stat_style'
        ws1['D' + str(row)].style = 'stat_style'
        ws1['E' + str(row)].style = 'stat_style'
        ws1['F' + str(row)].style = 'stat_style'
        ws1['G' + str(row)].style = 'stat_style'

        row += 1

        ws1['A' + str(row)] = "Mid"
        mid = results[name].starters[3]
        ws1['B' + str(row)] = mid.name
        ws1['C' + str(row)] = mid.k
        ws1['D' + str(row)] = mid.d
        ws1['E' + str(row)] = mid.a
        ws1['F' + str(row)] = mid.cs
        ws1['G' + str(row)] = mid.get_score()
        # apply formatting
        ws1['A' + str(row)].style = 'bold_style'
        ws1['B' + str(row)].style = 'base_style'
        ws1['C' + str(row)].style = 'stat_style'
        ws1['D' + str(row)].style = 'stat_style'
        ws1['E' + str(row)].style = 'stat_style'
        ws1['F' + str(row)].style = 'stat_style'
        ws1['G' + str(row)].style = 'stat_style'

        row += 1

        ws1['A' + str(row)] = "Bot"
        bot = results[name].starters[4]
        ws1['B' + str(row)] = bot.name
        ws1['C' + str(row)] = bot.k
        ws1['D' + str(row)] = bot.d
        ws1['E' + str(row)] = bot.a
        ws1['F' + str(row)] = bot.cs
        ws1['G' + str(row)] = bot.get_score()
        # apply formatting
        ws1['A' + str(row)].style = 'bold_style'
        ws1['B' + str(row)].style = 'base_style'
        ws1['C' + str(row)].style = 'stat_style'
        ws1['D' + str(row)].style = 'stat_style'
        ws1['E' + str(row)].style = 'stat_style'
        ws1['F' + str(row)].style = 'stat_style'
        ws1['G' + str(row)].style = 'stat_style'

        row += 1

        ws1['A' + str(row)] = "Support"
        supp = results[name].starters[0]
        ws1['B' + str(row)] = supp.name
        ws1['C' + str(row)] = supp.k
        ws1['D' + str(row)] = supp.d
        ws1['E' + str(row)] = supp.a
        ws1['F' + str(row)] = supp.cs
        ws1['G' + str(row)] = supp.get_score()
        # apply formatting
        ws1['A' + str(row)].style = 'bold_style'
        ws1['B' + str(row)].style = 'base_style'
        ws1['C' + str(row)].style = 'stat_style'
        ws1['D' + str(row)].style = 'stat_style'
        ws1['E' + str(row)].style = 'stat_style'
        ws1['F' + str(row)].style = 'stat_style'
        ws1['G' + str(row)].style = 'stat_style'

        row += 1

        # because sub dict uses weird entries, i think i need to loop through
        # should be inexpensive, just 2 per
        sub_num = 1
        for sub in results[name].subs:
            ws1['A' + str(row)] = "Sub-"+str(sub_num)
            ws1['B' + str(row)] = sub.name
            ws1['C' + str(row)] = sub.k
            ws1['D' + str(row)] = sub.d
            ws1['E' + str(row)] = sub.a
            ws1['F' + str(row)] = sub.cs
            ws1['G' + str(row)] = sub.get_score()
            # apply formatting
            ws1['A' + str(row)].style = 'bold_style'
            ws1['B' + str(row)].style = 'base_style'
            ws1['C' + str(row)].style = 'stat_style'
            ws1['D' + str(row)].style = 'stat_style'
            ws1['E' + str(row)].style = 'stat_style'
            ws1['F' + str(row)].style = 'stat_style'
            ws1['G' + str(row)].style = 'stat_style'

            row += 1
            sub_num += 1
        row += 1

    # make sure dest_filename is for a .xlsx file -- if other suffix is specified, it is ignored
    if not dest_filename.endswith('.xlsx'):
        # remove alternate suffix, if it is there
        temp = re.split('\.', dest_filename)
        dest_filename = temp[0] + '.xlsx' # get first half
    # wb.save(dest_filename)
    # os.startfile(dest_filename)

    # alternative: save in folder in project directory (or even outside)
    if not os.path.exists('fantasy_results'):
        os.makedirs('fantasy_results')
    final_path = 'fantasy_results\\' + dest_filename
    # final_path = os.path.expanduser('~\\Documents\\'+dest_filename)
    wb.save(final_path)
    os.startfile(final_path)

